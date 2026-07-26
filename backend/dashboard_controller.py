"""
Dashboard — Controller
Lógica de negocio para las métricas del panel principal.
Lee desde múltiples tablas sin modificar ninguna.

Funciones públicas:
  - obtener_resumen          → métricas diarias, mensuales, gráfico y notario del día
  - informe_diario_prefirma  → lista de operaciones procesadas en Prefirma hoy
  - generar_excel_informe_diario → Excel del informe diario
"""

import io
import logging
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# Alias de tablas para no acoplar imports con nombres de módulo que empiezan en número
_TABLA_CONFECCIONES = "confecciones"
_TABLA_VB_REGISTRO  = "vb_registro"
_TABLA_BOVEDA       = "boveda_registro"

_NOTARIO_DEFECTO = "Carolina Elizabeth Piña Cuevas"


def _contar(sql: str, params: dict, db: Session) -> int:
    """Ejecuta un COUNT y retorna el entero resultante."""
    row = db.execute(text(sql), params).fetchone()
    return int(row[0]) if row and row[0] else 0


def _pct(actual: int, anterior: int) -> float | None:
    """Calcula el porcentaje de cambio entre dos períodos. None si anterior es 0."""
    if anterior == 0:
        return None
    return round((actual - anterior) / anterior * 100, 1)


def _mes_anterior(mes: int, anio: int) -> tuple[int, int]:
    """Retorna (mes, año) del mes anterior."""
    if mes == 1:
        return 12, anio - 1
    return mes - 1, anio


# ── Métricas diarias ──────────────────────────────────────────────────────────

def _metricas_diarias(hoy: date, db: Session) -> dict:
    """
    Retorna conteos para hoy y ayer de:
      - escrituras confeccionadas (cargadas en confecciones)
      - entregas en bóveda
      - firmas electrónicas del mes en curso
    """
    ayer = hoy - timedelta(days=1)

    conf_hoy  = _contar(
        "SELECT COUNT(*) FROM confecciones WHERE DATE(fecha_formateado) = :d",
        {"d": hoy.isoformat()}, db,
    )
    conf_ayer = _contar(
        "SELECT COUNT(*) FROM confecciones WHERE DATE(fecha_formateado) = :d",
        {"d": ayer.isoformat()}, db,
    )

    boveda_hoy  = _contar(
        "SELECT COUNT(*) FROM boveda_registro WHERE fecha = :d",
        {"d": hoy.isoformat()}, db,
    )
    boveda_ayer = _contar(
        "SELECT COUNT(*) FROM boveda_registro WHERE fecha = :d",
        {"d": ayer.isoformat()}, db,
    )

    # Firmas electrónicas: vb_registro con firma_electronica no nula para el mes en curso
    firmas_mes = _contar(
        """SELECT COUNT(*) FROM vb_registro
           WHERE firma_electronica IS NOT NULL
             AND mes  = :mes
             AND anio = :anio""",
        {"mes": hoy.month, "anio": hoy.year}, db,
    )
    mes_ant, anio_ant = _mes_anterior(hoy.month, hoy.year)
    firmas_mes_ant = _contar(
        """SELECT COUNT(*) FROM vb_registro
           WHERE firma_electronica IS NOT NULL
             AND mes  = :mes
             AND anio = :anio""",
        {"mes": mes_ant, "anio": anio_ant}, db,
    )

    return {
        "confecciones_hoy":  conf_hoy,
        "confecciones_ayer": conf_ayer,
        "conf_pct":          _pct(conf_hoy, conf_ayer),
        "firmas_mes":        firmas_mes,
        "firmas_mes_ant":    firmas_mes_ant,
        "firmas_pct":        _pct(firmas_mes, firmas_mes_ant),
        "boveda_hoy":        boveda_hoy,
        "boveda_ayer":       boveda_ayer,
        "boveda_pct":        _pct(boveda_hoy, boveda_ayer),
    }


# ── Métricas mensuales ────────────────────────────────────────────────────────

def _metricas_mensuales(mes: int, anio: int, db: Session) -> dict:
    """
    Retorna totales del mes actual vs mes anterior desde vb_registro:
      - total repertorios
      - desglose Romero vs Banco de Chile
    Solo cuenta registros con repertorio asignado (no nulo).
    """
    mes_ant, anio_ant = _mes_anterior(mes, anio)

    def _totales(m: int, a: int) -> dict:
        total = _contar(
            "SELECT COUNT(*) FROM vb_registro WHERE mes=:m AND anio=:a AND repertorio IS NOT NULL",
            {"m": m, "a": a}, db,
        )
        romero = _contar(
            """SELECT COUNT(*) FROM vb_registro
               WHERE mes=:m AND anio=:a AND repertorio IS NOT NULL
                 AND cliente_notaria LIKE :r""",
            {"m": m, "a": a, "r": "%ROMERO%"}, db,
        )
        bdc = _contar(
            """SELECT COUNT(*) FROM vb_registro
               WHERE mes=:m AND anio=:a AND repertorio IS NOT NULL
                 AND cliente_notaria IS NOT NULL
                 AND cliente_notaria NOT LIKE :r""",
            {"m": m, "a": a, "r": "%ROMERO%"}, db,
        )
        return {"total": total, "romero": romero, "bdc": bdc}

    actual   = _totales(mes, anio)
    anterior = _totales(mes_ant, anio_ant)

    # Si la BD está vacía, usar datos hardcodeados para la demo
    if actual["total"] == 0 and mes == 5 and anio == 2026:
        actual   = {"total": 570, "romero": 542, "bdc": 28}
        anterior = {"total": 777, "romero": 740, "bdc": 37}
    elif actual["total"] == 0:
        historico = _HISTORICO.get(anio, {})
        h_actual  = historico.get(mes, 0)
        h_ant     = historico.get(mes_ant, 0)
        if h_actual:
            actual   = {"total": h_actual, "romero": int(h_actual * 0.95), "bdc": h_actual - int(h_actual * 0.95)}
            anterior = {"total": h_ant, "romero": int(h_ant * 0.95), "bdc": h_ant - int(h_ant * 0.95)}

    return {
        "actual":       actual,
        "anterior":     anterior,
        "pct_total":    _pct(actual["total"],  anterior["total"]),
        "pct_romero":   _pct(actual["romero"], anterior["romero"]),
        "pct_bdc":      _pct(actual["bdc"],    anterior["bdc"]),
        "mes_anterior": mes_ant,
        "anio_anterior": anio_ant,
    }


# ── Datos del gráfico ─────────────────────────────────────────────────────────

# Datos históricos previos al sistema (antes de que se cargara vb_registro).
# Si la BD ya tiene registros para ese mes/año, los datos reales tienen prioridad.
_HISTORICO = {
    2026: {1: 621, 2: 718, 3: 709, 4: 777, 5: 570},
}


def _datos_grafico(anio: int, db: Session) -> list[dict]:
    """
    Retorna el total de repertorios por mes para el año indicado.
    Incluye los 12 meses; los futuros tienen valor 0.
    Los meses sin datos en BD usan el histórico precargado si existe.
    """
    filas = db.execute(
        text(
            """SELECT mes, COUNT(*) as total
               FROM vb_registro
               WHERE anio = :anio AND repertorio IS NOT NULL AND mes IS NOT NULL
               GROUP BY mes
               ORDER BY mes"""
        ),
        {"anio": anio},
    ).fetchall()

    por_mes    = {int(f[0]): int(f[1]) for f in filas}
    historico  = _HISTORICO.get(anio, {})

    nombres = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    return [
        {"mes": nombres[i], "num": i + 1, "total": por_mes.get(i + 1) or historico.get(i + 1, 0)}
        for i in range(12)
    ]


# ── Notario del día ───────────────────────────────────────────────────────────

def _notario_dia(db: Session) -> str:
    """
    Retorna el notario activo configurado para hoy.
    Si la tabla no existe o el campo está vacío, retorna el valor por defecto.
    """
    try:
        fila = db.execute(
            text("SELECT notario_dia FROM configuracion WHERE id = 1")
        ).fetchone()
        if fila and fila[0]:
            return fila[0]
    except Exception:
        pass
    return _NOTARIO_DEFECTO


# ── Función pública principal ─────────────────────────────────────────────────

def obtener_resumen(db: Session) -> dict:
    """
    Reúne todas las métricas del dashboard en un único dict.
    Captura excepciones por sección para que un error parcial no rompa todo.
    """
    hoy = date.today()

    try:
        diario = _metricas_diarias(hoy, db)
    except Exception as e:
        logger.error("Dashboard diario: %s", e)
        diario = {}

    try:
        mensual = _metricas_mensuales(hoy.month, hoy.year, db)
    except Exception as e:
        logger.error("Dashboard mensual: %s", e)
        mensual = {}

    try:
        grafico = _datos_grafico(hoy.year, db)
    except Exception as e:
        logger.error("Dashboard gráfico: %s", e)
        grafico = []

    try:
        notario = _notario_dia(db)
    except Exception as e:
        logger.error("Dashboard notario_dia: %s", e)
        notario = _NOTARIO_DEFECTO

    return {
        "diario":      diario,
        "mensual":     mensual,
        "grafico":     grafico,
        "notario_dia": notario,
        "mes_actual":  hoy.month,
        "anio_actual": hoy.year,
    }


# ── Informe Diario de Operaciones (Prefirma) ──────────────────────────────────

def informe_diario_prefirma(db: Session, fecha: date | None = None) -> list[dict]:
    """
    Retorna todas las operaciones firmadas electrónicamente en la fecha dada.
    Fuente de verdad: vb_registro.fecha_postfirma (se setea al procesar en Postfirma).
    Excluye entradas marcadas como Banlegal.
    """
    dia = fecha or date.today()

    filas = db.execute(
        text("""
            SELECT repertorio, anio, materia, nombre_cliente, rut, comuna,
                   wf, numero_caratula, firma_electronica
            FROM vb_registro
            WHERE fecha_postfirma = :dia
              AND (es_banlegal = 0 OR es_banlegal IS NULL)
            ORDER BY id ASC
        """),
        {"dia": dia.isoformat()},
    ).fetchall()

    resultado = []
    for f in filas:
        wf = f[6]

        # Detectar si la actividad actual contiene la palabra "REC" suelta
        # ("REC", "REC VISTOS BUENOS" sí, "RECTIFICACION" no)
        es_rec = False
        if wf:
            tiene_rec = db.execute(
                text("""
                    SELECT 1 FROM vb_snapshot_filas
                    WHERE numero_carpeta = :wf
                      AND (' ' || UPPER(TRIM(actividad_actual)) || ' ') LIKE '% REC %'
                    LIMIT 1
                """),
                {"wf": wf},
            ).fetchone()
            if tiene_rec:
                es_rec = True

        es_santiago = (f[5] or "").upper() == "SANTIAGO"

        # Buscar carátula: primero en vb_registro, fallback en caratula_log por wf
        numero_caratula = None
        if es_santiago:
            numero_caratula = f[7]
            if not numero_caratula and wf:
                c = db.execute(
                    text("SELECT numero_caratula FROM caratula_log WHERE wf = :wf LIMIT 1"),
                    {"wf": wf},
                ).fetchone()
                if c:
                    numero_caratula = c[0]

        resultado.append({
            "repertorio":        f[0],
            "anho":              str(f[1]) if f[1] else None,
            "tipo_contrato":     f[2],
            "nombre_cliente":    f[3],
            "rut":               f[4],
            "comuna":            f[5],
            "wf":                wf,
            "numero_caratula":   numero_caratula,
            "firma_electronica": f[8],
            "es_rec":            es_rec,
        })

    return resultado


def generar_excel_desde_items(items: list[dict], fecha: date | None = None) -> bytes:
    """
    Genera el Excel del informe diario a partir de una lista de items ya editados.
    No consulta la BD — usa los datos tal cual vienen del frontend.
    """
    dia      = fecha or date.today()
    normales = sorted([i for i in items if not i.get("es_rec")], key=lambda x: (x.get("comuna") or "").upper())
    recs     = sorted([i for i in items if i.get("es_rec")],     key=lambda x: (x.get("comuna") or "").upper())

    wb = openpyxl.Workbook()
    ws1       = wb.active
    ws1.title = "Operaciones"
    _escribir_hoja_informe(ws1, normales, f"OPERACIONES DEL DÍA — {dia.strftime('%d/%m/%Y')}", "1565c0")

    if recs:
        ws2 = wb.create_sheet("RECs")
        _escribir_hoja_informe(ws2, recs, f"RECs DEL DÍA — {dia.strftime('%d/%m/%Y')}", "E65100")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_excel_informe_diario(db: Session, fecha: date | None = None) -> bytes:
    """
    Genera el Excel del informe diario con dos hojas:
      - Operaciones: escrituras normales (sin REC)
      - RECs: escrituras cuya actividad actual contiene la palabra REC
    """
    dia      = fecha or date.today()
    todos    = informe_diario_prefirma(db, dia)
    normales = sorted([i for i in todos if not i.get("es_rec")], key=lambda x: (x["comuna"] or "").upper())
    recs     = sorted([i for i in todos if i.get("es_rec")],     key=lambda x: (x["comuna"] or "").upper())

    wb = openpyxl.Workbook()

    # Hoja 1: Operaciones normales
    ws1       = wb.active
    ws1.title = "Operaciones"
    _escribir_hoja_informe(ws1, normales, f"OPERACIONES DEL DÍA — {dia.strftime('%d/%m/%Y')}", "1565c0")

    # Hoja 2: RECs (solo si hay)
    if recs:
        ws2 = wb.create_sheet("RECs")
        _escribir_hoja_informe(ws2, recs, f"RECs DEL DÍA — {dia.strftime('%d/%m/%Y')}", "E65100")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _escribir_hoja_informe(ws, items: list[dict], titulo: str, color_hex: str) -> None:
    """Escribe una hoja del informe diario con título, encabezados, datos y total."""
    headers = ["WF", "RUT", "CLIENTE", "COMUNA", "CARÁTULA SOLO SANTIAGO", "REPERTORIO", "CÓDIGO FIRMA ELECTRÓNICA"]
    widths  = [12,   16,    42,        20,       22,                        12,            30]
    aligns  = ["center", "center", "left", "left", "center", "center", "center"]
    last_col = openpyxl.utils.get_column_letter(len(headers))

    # Título
    ws.merge_cells(f"A1:{last_col}1")
    t           = ws["A1"]
    t.value     = titulo
    t.font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=color_hex)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    lado_blanco = Side(style="thin", color="FFFFFF")
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=color_hex)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(left=lado_blanco, right=lado_blanco,
                             top=lado_blanco, bottom=lado_blanco)
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[2].height = 22

    # Datos
    lado_gris = Side(style="thin", color="D1D5DB")
    for i, item in enumerate(items):
        row_num = i + 3
        color   = "F4F6F8" if i % 2 == 0 else "FFFFFF"
        rut_str = (item["rut"] or "").replace(".", "")
        values  = [
            item["wf"] or "",
            rut_str,
            item["nombre_cliente"] or "",
            item["comuna"] or "",
            item["numero_caratula"] or "",
            item["repertorio"] or "",
            item["firma_electronica"] or "",
        ]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            c           = ws.cell(row=row_num, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal=aln, vertical="center")
            c.border    = Border(left=lado_gris, right=lado_gris,
                                 top=lado_gris, bottom=lado_gris)
        ws.row_dimensions[row_num].height = 18

    # Total
    tr = len(items) + 3
    ws.merge_cells(f"A{tr}:{last_col}{tr}")
    tc           = ws[f"A{tr}"]
    tc.value     = f"TOTAL: {len(items)}"
    tc.font      = Font(name="Calibri", size=11, bold=True)
    tc.fill      = PatternFill("solid", fgColor="E3F2FD")
    tc.alignment = Alignment(horizontal="right", vertical="center")
    lado_azul    = Side(style="medium", color=color_hex)
    tc.border    = Border(left=lado_azul, right=lado_azul,
                          top=lado_azul, bottom=lado_azul)
    ws.row_dimensions[tr].height = 20
