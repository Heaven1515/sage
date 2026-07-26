"""
Módulo 02 — VB / Generación de Planillas
Controller: crea un archivo Excel por cada materia detectada, réplica exacta
de la planilla del sistema externo del Banco de Chile.

La planilla tiene:
  - Fila 1:  títulos de grupo (3 rangos fusionados + 1 celda suelta)
  - Fila 2:  nombres de columna individuales (altura 30, bordes bottom:thin)
  - Filas 3–36: datos (34 filas; BANCO DE CHILE pre-cargado en todas)

Los archivos se guardan en la carpeta del día activa (Tab 2).

Función pública:
  - generar_planillas(items, db) → retorna lista de nombres de archivos generados
"""

import logging
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .carpeta_model import VBCarpetaConfig
from .words_schema import WordItem

logger = logging.getLogger(__name__)

_RUT_BANCO      = "97004000-5"
_NOMBRE_BANCO   = "BANCO DE CHILE"
_CARPETA_MATRIZ = "Vistos Buenos de Abogados"

# ── Anchos exactos medidos en la planilla original ───────────────────────────
_ANCHOS_COLUMNA = {
    "A": 13.0,
    "B": 41.140625,
    "C": 20.5703125,
    "D": 33.5703125,
    "E": 13.0,
    "F": 41.140625,
    "G": 20.5703125,
    "H": 33.5703125,
    "I": 13.85546875,
    "J": 15.0,
    "K": 12.85546875,
    "U": 19.5703125,
}

# ── Helpers de estilo ─────────────────────────────────────────────────────────

_THIN = Side(border_style="thin")
_FUENTE = Font(name="Calibri", size=11)


def _borde(izquierda=False, derecha=False, arriba=False, abajo=False) -> Border:
    """Construye un Border thin solo en los lados indicados."""
    return Border(
        left=_THIN if izquierda else Side(),
        right=_THIN if derecha else Side(),
        top=_THIN if arriba else Side(),
        bottom=_THIN if abajo else Side(),
    )


# ── Construcción del libro ────────────────────────────────────────────────────

def _aplicar_anchos(ws) -> None:
    """Aplica los anchos de columna exactos de la planilla original."""
    for col_letra, ancho in _ANCHOS_COLUMNA.items():
        ws.column_dimensions[col_letra].width = ancho


def _construir_fila1(ws) -> None:
    """
    Fila 1: títulos de grupo.
    Tres rangos fusionados (A1:D1, E1:H1, I1:T1) y una celda suelta (U1).
    Bordes izquierdo y derecho en cada bloque; texto centrado.
    """
    grupos = [
        ("A1:D1", "A1", "D1", "Compareciente 1 / Vendedor"),
        ("E1:H1", "E1", "H1", "Compareciente 2 / Comprador"),
        ("I1:T1", "I1", "T1", "Vehículo"),
    ]
    for rango, inicio, fin, titulo in grupos:
        ws.merge_cells(rango)
        celda_inicio = ws[inicio]
        celda_inicio.value = titulo
        celda_inicio.font = _FUENTE
        celda_inicio.alignment = Alignment(horizontal="center")
        celda_inicio.border = _borde(izquierda=True, derecha=False)
        # Borde derecho en la última celda del grupo (no fusionada con la primera)
        ws[fin].border = _borde(derecha=True)

    # U1: celda suelta — solo borde izquierdo y derecho
    u1 = ws["U1"]
    u1.value = "Datos OT"
    u1.font = _FUENTE
    u1.border = _borde(izquierda=True, derecha=True)


def _construir_fila2(ws) -> None:
    """
    Fila 2: nombres de columna individuales.
    Altura 30. Todas con borde bottom:thin.
    wrap_text=True solo en B2 y F2 (texto multilínea en la planilla original).
    Bordes adicionales: left en A2 y U2; right en D2, H2, T2, U2.
    """
    ws.row_dimensions[2].height = 30.0

    columnas = [
        # (letra, valor, wrap, izquierda, derecha)
        ("A", "RUT",                              False, True,  False),
        ("B", "Apellido Paterno /\nRazón Social", True,  False, False),
        ("C", "Apellido Materno",                 False, False, False),
        ("D", "Nombres",                          False, False, True),
        ("E", "RUT",                              False, False, False),
        ("F", "Apellido Paterno /\nRazón Social", True,  False, False),
        ("G", "Apellido Materno",                 False, False, False),
        ("H", "Nombres",                          False, False, True),
        ("I", "Tasación Fiscal",                  False, False, False),
        ("J", "Precio de Venta",                  False, False, False),
        ("K", "Patente",                          False, False, False),
        ("L", "Tipo",                             False, False, False),
        ("M", "Marca",                            False, False, False),
        ("N", "Modelo",                           False, False, False),
        ("O", "Año",                              False, False, False),
        ("P", "Color",                            False, False, False),
        ("Q", "Motor",                            False, False, False),
        ("R", "Chasis",                           False, False, False),
        ("S", "Serie",                            False, False, False),
        ("T", "Vin",                              False, False, True),
        ("U", "Código de Operación",              False, True,  True),
    ]
    for letra, valor, wrap, izq, der in columnas:
        celda = ws[f"{letra}2"]
        celda.value = valor
        celda.font = _FUENTE
        celda.alignment = Alignment(wrap_text=wrap if wrap else None)
        celda.border = _borde(izquierda=izq, derecha=der, abajo=True)


def _construir_datos(ws, items: list[WordItem]) -> None:
    """
    Filas desde la 3 en adelante: todas las que sean necesarias, sin límite.
    Todas llevan BANCO DE CHILE pre-cargado en A y B.
    Cada item lleva: RUT en E, nombre en F, WF (entero) en U.
    Sin bordes ni formato especial en filas de datos (igual que el original).
    """
    for i, item in enumerate(items):
        fila = 3 + i
        ws.cell(row=fila, column=1, value=_RUT_BANCO).font = _FUENTE
        ws.cell(row=fila, column=2, value=_NOMBRE_BANCO).font = _FUENTE
        ws.cell(row=fila, column=5, value=item.rut or "").font = _FUENTE
        ws.cell(row=fila, column=6, value=item.nombre_cliente).font = _FUENTE
        # WF como entero si es numérico (igual que la planilla original)
        try:
            wf_valor = int(item.wf)
        except (ValueError, TypeError):
            wf_valor = item.wf
        ws.cell(row=fila, column=21, value=wf_valor).font = _FUENTE


def _crear_libro(items: list[WordItem]) -> openpyxl.Workbook:
    """Ensambla el Workbook completo con la estructura y formato exactos del original."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    _aplicar_anchos(ws)
    _construir_fila1(ws)
    _construir_fila2(ws)
    _construir_datos(ws, items)

    return wb


# ── Helpers de ruta ───────────────────────────────────────────────────────────

def _obtener_ruta_carpeta_dia(db: Session) -> Path:
    """
    Obtiene la ruta de la carpeta del día activa desde la BD.
    Lanza ValueError si la carpeta no está configurada o no se ha activado hoy.
    """
    config = db.query(VBCarpetaConfig).filter(VBCarpetaConfig.id == 1).first()
    if not config or not config.ruta_base or not config.carpeta_hoy:
        raise ValueError(
            "No hay carpeta del día configurada. "
            "Activa el vigilador en la sección Carpeta de Descargas VB."
        )
    ruta = Path(config.ruta_base) / _CARPETA_MATRIZ / config.carpeta_hoy
    if not ruta.exists():
        raise ValueError(
            f"La carpeta del día '{config.carpeta_hoy}' no existe en disco. "
            "Vuelve a activar el vigilador para recrearla."
        )
    return ruta


def _sin_tildes(texto: str) -> str:
    """Elimina tildes y acentos. 'PROHIBICIÓN' → 'PROHIBICION'."""
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def _normalizar_materia(materia: str) -> str:
    """Normaliza la materia quitando tildes y pasando a mayúsculas."""
    return _sin_tildes(materia.strip().upper())


def _nombre_archivo(materia: str) -> str:
    """
    Construye el nombre del archivo Excel.
    Ej: "ALZAMIENTO DE HIPOTECA" → PLANILLA_ALZAMIENTO_DE_HIPOTECA_03-04-2026.xlsx
    """
    materia_limpia = _normalizar_materia(materia).replace(" ", "_")
    fecha = date.today().strftime("%d-%m-%Y")
    return f"PLANILLA_{materia_limpia}_{fecha}.xlsx"


# ── Función pública ───────────────────────────────────────────────────────────

def generar_planillas(items: list[WordItem], db: Session) -> list[str]:
    """
    Agrupa los items por materia y genera un Excel por cada grupo.
    Sin límite de filas — todos los items del grupo van en una sola planilla.
    Los archivos se guardan en la carpeta del día activa (Tab 2).
    Retorna la lista de nombres de archivos generados.
    """
    if not items:
        raise ValueError("No hay documentos cargados para generar planillas")

    ruta_carpeta = _obtener_ruta_carpeta_dia(db)

    # Agrupar preservando el orden de aparición (normaliza tildes para no separar
    # "PROHIBICION" de "PROHIBICIÓN" en planillas distintas)
    por_materia: dict[str, list[WordItem]] = {}
    for item in items:
        materia = _normalizar_materia(item.materia or "SIN_MATERIA")
        por_materia.setdefault(materia, []).append(item)

    archivos_generados: list[str] = []

    for materia, grupo in por_materia.items():
        nombre = _nombre_archivo(materia)
        ruta = ruta_carpeta / nombre
        try:
            libro = _crear_libro(grupo)
            libro.save(str(ruta))
            archivos_generados.append(nombre)
            logger.info(
                "Planilla generada: %s (%d registros) → %s",
                nombre, len(grupo), ruta,
            )
        except Exception as e:
            logger.error("Error guardando planilla '%s': %s", nombre, e)
            raise ValueError(
                f"No se pudo generar la planilla para '{materia}': {e}"
            ) from e

    return archivos_generados
